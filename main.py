import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Import custom modules
from src.Utils.K_Fold import K_Fold
from src.data_loader import load_data
from src.model.train_model import get_X_y, train_model
from src.analysis.SHAP import shap_analysis
from src.analysis.LIME import lime_sensitivity_analysis
from src.Optimiser.HOA.hoa_optimizer import hoa_optimizer
from src.Utils.balance_dataset_smote_enc import balance_with_smote_enc
from objective_function import objective_stacking

# =======================
# 📌 Configuration
# =======================
DATA_PATH = (
    r"D:\ML\B(SMOTE-ENC)#K-FOLD3#M(XGBC&RFC)#O(HOA)#A(SHAP[PLOTS])\data\data.xlsx"
)
RAW_CSV_PATH = r"D:\ML\B(SMOTE-ENC)#K-FOLD3#M(XGBC&RFC)#O(HOA)#A(SHAP[PLOTS])\data\Dataset-10MO-Yahyavi (Student Stress Level).csv"
TARGET_COLUMN = "Stress Level "
CATEGORICAL_FEATURES = ["Gender"]

# =======================
# 📥 Data Loading & Balancing
# =======================

# Load raw data from Excel
raw_data = load_data(DATA_PATH)

# Balance dataset using SMOTE-ENC
balanced_df = balance_with_smote_enc(
    file_path=RAW_CSV_PATH,
    target_col=TARGET_COLUMN,
    categorical_cols=CATEGORICAL_FEATURES,
)

# Encode 'Gender' column
balanced_df["Gender"] = balanced_df["Gender"].map({"Male": 0, "Female": 1, "Other": 2})

# Round numerical values
balanced_df = balanced_df.round(1)

# # Save balanced data to Excel (overwrite existing sheet if needed)
# book = load_workbook(DATA_PATH)
# if "Data after SMOTE-ENC" in book.sheetnames:
#     book.remove(book["Data after SMOTE-ENC"])
#     book.save(DATA_PATH)

# with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
#     balanced_df.to_excel(writer, sheet_name="Data after SMOTE-ENC", index=False)

# =======================
# ✂️ Feature Extraction
# =======================
X, y = get_X_y(balanced_df, target_col=TARGET_COLUMN)

# =======================
# 🔁 K-Fold Cross-Validation
# =======================
X_train, X_test, y_train, y_test, kfold_scores, combined_df = K_Fold(X, y, n_splits=5)
kfold_scores_df = pd.DataFrame(kfold_scores)

# # Save combined K-Fold dataset to Excel
# book = load_workbook(DATA_PATH)
# if "DATA after K-Fold" in book.sheetnames:
#     book.remove(book["DATA after K-Fold"])
#     book.save(DATA_PATH)

# with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a") as writer:
#     combined_df.to_excel(writer, sheet_name="DATA after K-Fold", index=False)

# =======================
# ⚙️ Baseline Model Training (Without Optimization)
# =======================
baseline_result = train_model(X_train, y_train, X_test, y_test)


# # =======================
# # 🚀 Run HOA Optimizer on Objective Function
# # =======================
# best_params, best_rmse, convergence = hoa_optimizer(
#     objective_function=objective_stacking,
#     lb=[50, 0.01],  # Lower bounds for [n_estimators, learning_rate]
#     ub=[300, 1.0],  # Upper bounds
#     dim=2,  # Number of parameters
#     n_agents=4,  # Number of agents (keep low for testing)
#     max_iter=10,  # Number of iterations (increase for better optimization)
#     X_train=X_train,
#     y_train=y_train,
#     X_test=X_test,
#     y_test=y_test,
# )

# # =======================
# # 🤖 Retrain Model with Optimized Parameters
# # =======================
# hoa_result = train_model(X_train, y_train, X_test, y_test, params=best_params)

# # =======================
# # 📊 SHAP Analysis
# # =======================
# shap_df, shap_values = shap_analysis(
#     model=hoa_result["model"],
#     X_train=X_train,
#     y_train=y_train,
#     X_test=X_test,
#     y_test=y_test,
#     save_path=DATA_PATH,
#     sheet_name="SHAP_Sensitivity",
# )

# # =======================
# # 🧪 LIME Sensitivity Analysis
# # =======================
# lime_result = lime_sensitivity_analysis(
#     model=hoa_result["model"],
#     X_train=X_train,
#     y_train=y_train,
#     X_test=X_test,
#     y_test=y_test,
#     sample_index=5,
#     epsilon=0.05,
# )

# # =======================
# # 📤 Final Output
# # =======================
# print("\n✅ Best Hyperparameters (from HOA):", best_params)
# print("📉 Best RMSE:", best_rmse)
